#!python3
'''
    updateProduce.py - corrige os preços em uma planilha de venda de produtos
'''
import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
PRICE_UPDATES = {'Garlic': 4.01, 'Celery': 2.19, 'Lemon': 0.99}
# Percorre todas as linhas em um loop e atualiza os preços
for rowNum in range(2, len(sheet['A'])):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedSales.xlsx')
