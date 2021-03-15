#!python3
'''
    sendDuesReminders.py - Envia emails de acordo com o status de pagamento na planilha
'''
import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
unpaidMembers = {}
lastRow = sheet.max_row
for r in range(2, lastRow + 1):
    payment = sheet.cell(row = r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name]= email

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('gtestes2020@gmail.com', 'czlgxpkugmcfpkqj')

for name, email in unpaidMembers.items():
    body = f'Subject:{latestMonth} dues unpaid.\nDear {name},\nRecords show that you have not paid dues for {latestMonth}. Plesae make this payment as soon as possible. Thank you!'
    print(f'Sending email to {email}...')
    sendmailStatus = smtpObj.sendmail('gtestes2020@gmail.com', email, body)
    if sendmailStatus != {}:
        print(f'There was a problem sending email to {email}: {sendmailStatus}.')
smtpObj.quit()